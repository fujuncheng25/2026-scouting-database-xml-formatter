import io
import os
import secrets
from datetime import datetime, timezone
from functools import wraps
from xml.etree.ElementTree import Element, SubElement, tostring

import psycopg2
from psycopg2.extras import Json, RealDictCursor
from flask import Flask, Response, jsonify, render_template, request, url_for
from openpyxl import Workbook
from werkzeug.security import check_password_hash, generate_password_hash
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.config["JSON_SORT_KEYS"] = False

MATCH_TYPE_MAP = {
    "P": "Practice",
    "Q": "Qualification",
    "M": "Match",
    "F": "Final"
}

FIELD_ROWS = [
    ("Autonomous","Shooter Type","autonomousShootertype"),
    ("Autonomous","Fuel Out","autonomousShotstaken"),
    ("Autonomous","Fuel In","autonomousShotvolumes"),
    ("Teleop","Fuel Out","teleopFuelcount"),
    ("Teleop","Fuel In","teleopShotstaken"),
    ("Teleop","Human Fuel Count","teleopHumanfuelcount"),
    ("Teleop","Pass Bump","teleopPassbump"),
    ("Teleop","Pass Trench","teleopPasstrench"),
    ("Teleop","Fetch Ball Preference","teleopFetchballpreference"),  # 新增字段
    ("End&AfterGame","Tower Status","endAndAfterGameTowerstatus"),
    ("End&AfterGame","Climbing Time","endAndAfterGameClimbingtime"),
    ("End&AfterGame","Ranking Points","endAndAfterGameRankingpoint"),
    ("End&AfterGame","Coop Point","endAndAfterGameCooppoint"),
    ("End&AfterGame","Autonomous Move","endAndAfterGameAutonomousmove"),
    ("End&AfterGame","Teleop Move","endAndAfterGameTeleopmove"),
    ("End&AfterGame","Comments","endAndAfterGameComments"),
]

def now_utc():
    return datetime.now(timezone.utc)


def get_remote_db():
    return psycopg2.connect(
        host=os.getenv("REMOTE_POSTGRES_HOST", os.getenv("POSTGRES_HOST", "localhost")),
        port=os.getenv("REMOTE_POSTGRES_PORT", os.getenv("POSTGRES_PORT", "5432")),
        user=os.getenv("REMOTE_POSTGRES_USER", os.getenv("POSTGRES_USER", "postgres")),
        password=os.getenv("REMOTE_POSTGRES_PASSWORD", os.getenv("POSTGRES_PASSWORD", "postgres")),
        dbname=os.getenv("REMOTE_POSTGRES_DB", os.getenv("POSTGRES_DB", "postgres")),
    )


def get_local_db():
    return psycopg2.connect(
        host=os.getenv("LOCAL_POSTGRES_HOST", "localhost"),
        port=os.getenv("LOCAL_POSTGRES_PORT", "5432"),
        user=os.getenv("LOCAL_POSTGRES_USER", "postgres"),
        password=os.getenv("LOCAL_POSTGRES_PASSWORD", "postgres"),
        dbname=os.getenv("LOCAL_POSTGRES_DB", "scouting_local"),
    )


def ensure_local_schema():
    conn = get_local_db()
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS local_snapshots (
                remote_record_id BIGINT PRIMARY KEY,
                scout_event_id TEXT NOT NULL,
                match_type_code TEXT NOT NULL,
                match_number INTEGER,
                team_number TEXT,
                alliance TEXT,
                payload JSONB NOT NULL,
                remote_updated_at TIMESTAMPTZ NOT NULL,
                synced_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_local_snapshots_lookup
            ON local_snapshots (scout_event_id, match_type_code, match_number);
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS local_edits (
                id BIGSERIAL PRIMARY KEY,
                remote_record_id BIGINT NOT NULL REFERENCES local_snapshots(remote_record_id) ON DELETE CASCADE,
                field_key TEXT NOT NULL,
                field_value TEXT,
                editor TEXT NOT NULL DEFAULT 'anonymous',
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                UNIQUE (remote_record_id, field_key)
            );
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_local_edits_lookup
            ON local_edits (remote_record_id, field_key);
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS local_users (
                id BIGSERIAL PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                active BOOLEAN NOT NULL DEFAULT TRUE
            );
            """
        )
    conn.commit()
    conn.close()


def normalize_match_code(match_type):
    if not match_type:
        return None
    code = str(match_type).strip().upper()[:1]
    return code if code in MATCH_TYPE_MAP else None


def key_candidates(key):
    snake = []
    for ch in key:
        if ch.isupper():
            snake.extend(["_", ch.lower()])
        else:
            snake.append(ch)
    snake_key = "".join(snake)
    return [key, snake_key]

def format_value(v):
    if v is None:
        return ""
    if isinstance(v,bool):
        return "Yes" if v else "No"
    return str(v)

def get_value(record, key):
    for candidate in key_candidates(key):
        if candidate in record:
            return record.get(candidate)
    return None


def get_remote_value(record, *keys):
    for key in keys:
        if key in record:
            return record[key]
    return None


def auth_required():
    return os.getenv("REQUIRE_AUTH", "false").strip().lower() in {"1", "true", "yes", "on"}


def get_current_user():
    username = (request.args.get("u") or request.headers.get("X-Auth-User") or "").strip()
    password = request.args.get("p") or request.headers.get("X-Auth-Pass")
    if not auth_required():
        return username or "anonymous"
    if not username or not password:
        return None
    conn = get_local_db()
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(
            """
            SELECT username, password_hash, active
            FROM local_users
            WHERE username = %s
            """,
            (username,),
        )
        user = cur.fetchone()
    conn.close()
    if not user or not user["active"]:
        return None
    if not check_password_hash(user["password_hash"], password):
        return None
    return user["username"]


def require_optional_auth(fn):
    @wraps(fn)
    def wrapped(*args, **kwargs):
        user = get_current_user()
        if auth_required() and user is None:
            if request.path.startswith("/api") or request.path.startswith("/auth"):
                return jsonify({"error": "Authentication required. Use ?u=<username>&p=<password>."}), 401
            return Response("Authentication required. Use URL params u and p.", status=401)
        request.current_actor = user or "anonymous"
        return fn(*args, **kwargs)

    return wrapped


def load_remote_rows(event_id, match_type_code, match_number=None):
    db_type = MATCH_TYPE_MAP.get(match_type_code)
    if not db_type:
        return []
    conn = get_remote_db()
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        if match_number is None:
            cur.execute(
                """
                SELECT *
                FROM team_match_record
                WHERE "scoutEventId" = %s
                  AND "matchType" = %s
                ORDER BY "id" ASC
                """,
                (event_id, db_type),
            )
        else:
            cur.execute(
                """
                SELECT *
                FROM team_match_record
                WHERE "scoutEventId" = %s
                  AND "matchType" = %s
                  AND "matchNumber" = %s
                ORDER BY "id" ASC
                """,
                (event_id, db_type, match_number),
            )
        rows = cur.fetchall()
    conn.close()
    return rows


def sync_remote_to_local(event_id, match_type_code, match_number=None):
    rows = load_remote_rows(event_id, match_type_code, match_number)
    conn = get_local_db()
    synced_count = 0
    changed_count = 0
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        for row in rows:
            remote_record_id = get_remote_value(row, "id")
            if remote_record_id is None:
                continue

            current_payload = dict(row)
            source_updated_at = get_remote_value(row, "updatedAt", "updated_at", "modifiedAt", "modified_at")

            cur.execute(
                """
                SELECT payload, remote_updated_at
                FROM local_snapshots
                WHERE remote_record_id = %s
                """,
                (remote_record_id,),
            )
            existing = cur.fetchone()
            payload_changed = True
            if existing:
                payload_changed = existing["payload"] != current_payload

            if source_updated_at:
                effective_remote_updated_at = source_updated_at
            elif existing and not payload_changed:
                effective_remote_updated_at = existing["remote_updated_at"]
            elif existing and payload_changed:
                effective_remote_updated_at = now_utc()
                changed_count += 1
            else:
                effective_remote_updated_at = now_utc()
                changed_count += 1

            cur.execute(
                """
                INSERT INTO local_snapshots (
                    remote_record_id,
                    scout_event_id,
                    match_type_code,
                    match_number,
                    team_number,
                    alliance,
                    payload,
                    remote_updated_at,
                    synced_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT (remote_record_id)
                DO UPDATE SET
                    scout_event_id = EXCLUDED.scout_event_id,
                    match_type_code = EXCLUDED.match_type_code,
                    match_number = EXCLUDED.match_number,
                    team_number = EXCLUDED.team_number,
                    alliance = EXCLUDED.alliance,
                    payload = EXCLUDED.payload,
                    remote_updated_at = EXCLUDED.remote_updated_at,
                    synced_at = NOW();
                """,
                (
                    remote_record_id,
                    str(get_remote_value(row, "scoutEventId", "scout_event_id") or event_id),
                    match_type_code,
                    get_remote_value(row, "matchNumber", "match_number"),
                    str(get_remote_value(row, "teamNumber", "team_number") or ""),
                    str(get_remote_value(row, "alliance") or ""),
                    Json(current_payload),
                    effective_remote_updated_at,
                ),
            )
            synced_count += 1
    conn.commit()
    conn.close()
    return {"synced": synced_count, "updated_remote_records": changed_count}


def fetch_effective_matches(event_id, match_type_code, match_number=None):
    conn = get_local_db()
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        if match_number is None:
            cur.execute(
                """
                SELECT *
                FROM local_snapshots
                WHERE scout_event_id = %s
                  AND match_type_code = %s
                ORDER BY match_number NULLS LAST, team_number, remote_record_id
                """,
                (event_id, match_type_code),
            )
        else:
            cur.execute(
                """
                SELECT *
                FROM local_snapshots
                WHERE scout_event_id = %s
                  AND match_type_code = %s
                  AND match_number = %s
                ORDER BY team_number, remote_record_id
                """,
                (event_id, match_type_code, match_number),
            )
        snapshots = cur.fetchall()

        record_ids = [row["remote_record_id"] for row in snapshots]
        edits = {}
        if record_ids:
            cur.execute(
                """
                SELECT remote_record_id, field_key, field_value, updated_at
                FROM local_edits
                WHERE remote_record_id = ANY(%s::bigint[])
                ORDER BY updated_at DESC
                """,
                (record_ids,),
            )
            for row in cur.fetchall():
                per_record = edits.setdefault(row["remote_record_id"], {})
                existing = per_record.get(row["field_key"])
                if existing is None or row["updated_at"] > existing["updated_at"]:
                    per_record[row["field_key"]] = {
                        "value": row["field_value"],
                        "updated_at": row["updated_at"],
                    }
    conn.close()

    matches = {}
    for snapshot in snapshots:
        payload = dict(snapshot["payload"])
        per_record_edits = edits.get(snapshot["remote_record_id"], {})
        for field_key, edit in per_record_edits.items():
            if edit["updated_at"] >= snapshot["remote_updated_at"]:
                payload[field_key] = edit["value"]

        record = {
            "id": snapshot["remote_record_id"],
            "teamNumber": snapshot["team_number"],
            "alliance": snapshot["alliance"],
            "payload": payload,
        }
        current_match_number = snapshot["match_number"]
        match = matches.setdefault(current_match_number, {"red": [], "blue": []})
        if "red" in str(snapshot["alliance"]).lower():
            match["red"].append(record)
        else:
            match["blue"].append(record)
    return dict(sorted(matches.items(), key=lambda item: (item[0] is None, item[0]))), len(snapshots)

def build_matrix(match):
    columns = []
    for t in match["red"]:
        columns.append({"label": f"Red {t['teamNumber']}", "record": t})
    for t in match["blue"]:
        columns.append({"label": f"Blue {t['teamNumber']}", "record": t})
    while len(columns) < 6:
        if len(columns) < 3:
            columns.append({"label": "Red", "record": None})
        else:
            columns.append({"label": "Blue", "record": None})

    rows = []
    for cat, name, key in FIELD_ROWS:
        cells = []
        for c in columns:
            if c["record"]:
                payload = c["record"]["payload"]
                v = get_value(payload, key)
                cells.append(
                    {
                        "value": format_value(v),
                        "record_id": c["record"]["id"],
                        "field_key": key,
                        "editable": True,
                    }
                )
            else:
                cells.append({"value": "", "editable": False})
        rows.append({"section": cat, "field": name, "field_key": key, "cells": cells})
    return columns, rows

def fill_sheet(ws, match):
    columns, rows = build_matrix(match)
    ws.cell(row=1, column=1, value="Section")
    ws.cell(row=1, column=2, value="Field")
    for i, c in enumerate(columns, start=3):
        ws.cell(row=1, column=i, value=c["label"])
    for r_i, row in enumerate(rows, start=2):
        flattened = [row["section"], row["field"]] + [cell["value"] for cell in row["cells"]]
        for c_i, val in enumerate(flattened, start=1):
            ws.cell(row=r_i, column=c_i, value=val)

def create_excel(matches, match_type_code):
    wb = Workbook()
    sheet_prefix = (match_type_code or "M").strip().upper()[:1] or "M"
    if not matches:
        ws = wb.active
        ws.title = f"{sheet_prefix}NoData"
        ws.cell(row=1, column=1, value="No data found")
    else:
        sorted_items = sorted(matches.items(), key=lambda item: (item[0] is None, item[0]))
        first_match_number, first_match = sorted_items[0]
        first_title = f"{sheet_prefix}Unknown" if first_match_number is None else f"{sheet_prefix}{first_match_number}"
        ws = wb.active
        ws.title = first_title[:31]
        fill_sheet(ws, first_match)
        for match_number, match in sorted_items[1:]:
            title = f"{sheet_prefix}Unknown" if match_number is None else f"{sheet_prefix}{match_number}"
            page = wb.create_sheet(title=title[:31])
            fill_sheet(page, match)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def create_xml(match):
    root = Element("TeamMatchRecord")
    columns, rows = build_matrix(match)
    headers = SubElement(root, "Headers")
    SubElement(headers, "Header").text = "Section"
    SubElement(headers, "Header").text = "Field"
    for c in columns:
        SubElement(headers, "Header").text = c["label"]
    body = SubElement(root, "Rows")
    for row in rows:
        r = SubElement(body, "Row")
        flattened = [row["section"], row["field"]] + [cell["value"] for cell in row["cells"]]
        for cell in flattened:
            SubElement(r, "Cell").text = str(cell)
    return tostring(root, encoding="utf-8", xml_declaration=True)


def build_pages(matches):
    pages = []
    for current_match_number, match in matches.items():
        columns, rows = build_matrix(match)
        pages.append(
            {
                "match_number": current_match_number,
                "columns": columns,
                "rows": rows,
            }
        )
    return pages

@app.route("/")
@require_optional_auth
def index():
    ensure_local_schema()
    event_id = request.args.get("scout_event_id")
    match_type = normalize_match_code(request.args.get("match_type"))
    match_number_raw = (request.args.get("match_number") or "").strip()
    match_number = int(match_number_raw) if match_number_raw.isdigit() else None
    pages = []
    sync_report = None
    snapshot_count = 0
    auth_user_param = request.args.get("u", "")
    auth_pass_param = request.args.get("p", "")
    auth_query = ""
    if auth_user_param and auth_pass_param:
        auth_query = f"&u={auth_user_param}&p={auth_pass_param}"

    if event_id and match_type:
        sync_report = sync_remote_to_local(event_id, match_type, match_number)
        matches, snapshot_count = fetch_effective_matches(event_id, match_type, match_number)
        pages = build_pages(matches)

    return render_template(
        "index.html",
        pages=pages,
        scout_event_id=event_id,
        match_type=match_type,
        match_number=match_number_raw,
        snapshot_count=snapshot_count,
        sync_report=sync_report,
        require_auth=auth_required(),
        actor=getattr(request, "current_actor", "anonymous"),
        auth_user_param=auth_user_param,
        auth_pass_param=auth_pass_param,
        auth_query=auth_query,
    )

@app.get("/export/excel")
@require_optional_auth
def export_excel():
    ensure_local_schema()
    event_id = request.args.get("scout_event_id")
    match_type = normalize_match_code(request.args.get("match_type"))
    match_number_raw = (request.args.get("match_number") or "").strip()
    match_number = int(match_number_raw) if match_number_raw.isdigit() else None
    if event_id and match_type:
        sync_remote_to_local(event_id, match_type, match_number)
    matches, _ = fetch_effective_matches(event_id, match_type, match_number)
    data = create_excel(matches, match_type)
    if match_number is None:
        name = f"matches_{datetime.now().strftime('%Y%m%d')}.xlsx"
    else:
        name = f"match_{match_number}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )

@app.get("/export/xml")
@require_optional_auth
def export_xml():
    ensure_local_schema()
    event_id = request.args.get("scout_event_id") or request.args.get("event_id")
    match_type = normalize_match_code(request.args.get("match_type"))
    match_number = int(request.args.get("match_number"))
    sync_remote_to_local(event_id, match_type, match_number)
    matches, _ = fetch_effective_matches(event_id, match_type, match_number)
    match = matches.get(match_number, {"red": [], "blue": []})
    data = create_xml(match)
    name = f"match_{match_number}_{datetime.now().strftime('%Y%m%d')}.xml"
    return Response(
        data,
        mimetype="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


@app.post("/api/sync")
@require_optional_auth
def api_sync():
    ensure_local_schema()
    payload = request.get_json(silent=True) or {}
    event_id = payload.get("scout_event_id") or request.args.get("scout_event_id")
    match_type = normalize_match_code(payload.get("match_type") or request.args.get("match_type"))
    match_number_raw = payload.get("match_number")
    if match_number_raw in (None, ""):
        match_number_raw = request.args.get("match_number")
    match_number = int(match_number_raw) if str(match_number_raw).isdigit() else None

    if not event_id or not match_type:
        return jsonify({"error": "scout_event_id and match_type are required"}), 400

    report = sync_remote_to_local(event_id, match_type, match_number)
    return jsonify({"ok": True, "report": report})


@app.post("/api/edit")
@require_optional_auth
def api_edit():
    ensure_local_schema()
    payload = request.get_json(silent=True) or {}
    record_id = payload.get("record_id")
    field_key = (payload.get("field_key") or "").strip()
    field_value = payload.get("field_value", "")

    if not record_id or not field_key:
        return jsonify({"error": "record_id and field_key are required"}), 400

    conn = get_local_db()
    with conn.cursor() as cur:
        cur.execute("SELECT 1 FROM local_snapshots WHERE remote_record_id = %s", (record_id,))
        exists = cur.fetchone()
        if not exists:
            conn.close()
            return jsonify({"error": "Record is not present in local cache. Run sync first."}), 404

        cur.execute(
            """
            INSERT INTO local_edits (remote_record_id, field_key, field_value, editor, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            ON CONFLICT (remote_record_id, field_key)
            DO UPDATE SET
                field_value = EXCLUDED.field_value,
                editor = EXCLUDED.editor,
                updated_at = NOW();
            """,
            (record_id, field_key, str(field_value), getattr(request, "current_actor", "anonymous")),
        )
    conn.commit()
    conn.close()

    return jsonify({"ok": True, "record_id": record_id, "field_key": field_key, "saved_at": now_utc().isoformat()})


@app.post("/auth/bootstrap")
def auth_bootstrap_user():
    ensure_local_schema()
    username = f"viewer_{secrets.token_hex(3)}"
    password = secrets.token_urlsafe(10)
    conn = get_local_db()
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO local_users (username, password_hash)
            VALUES (%s, %s)
            """,
            (username, generate_password_hash(password)),
        )
    conn.commit()
    conn.close()

    access_url = f"{request.host_url.rstrip('/')}{url_for('index')}?u={username}&p={password}"
    return jsonify(
        {
            "ok": True,
            "username": username,
            "password": password,
            "access_url": access_url,
            "auth_required": auth_required(),
            "note": "If REQUIRE_AUTH=false this account is optional, but still available for collaboration links.",
        }
    )


@app.get("/healthz")
def healthz():
    ensure_local_schema()
    return jsonify({"ok": True})

if __name__ == "__main__":
    ensure_local_schema()
    app.run(host="0.0.0.0", port=5000, debug=True)